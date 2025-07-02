import excel_parsing as ep
import sys
from PyQt6.QtWidgets import (QApplication, QMainWindow, QPushButton, QLabel, 
                             QDialog, QVBoxLayout,
                             QTableWidget, QTableWidgetItem, QDialogButtonBox, QFileDialog, QFrame,QComboBox)
from PyQt6.QtCore import Qt
from datetime import datetime
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
from docx import Document
import word_parsing as wp

MONTHS_RU = {
    1: "январь", 2: "февраль", 3: "март", 4: "апрель",
    5: "май", 6: "июнь", 7: "июль", 8: "август",
    9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь"
}

work_list_filepath = "work_list.xlsx" # Имя Ecxel файла с перечисленными работами
work_list = []
dds_list = []
selected_work = ""

sub_map = { } # Хранит параграф и позицию найденных подстрок в документе

class ShopConfirmationDialog(QDialog):
    def __init__(self, file_shop_map, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Подтверждение магазинов")
        self.setGeometry(200, 200, 600, 400)
        
        layout = QVBoxLayout()
        
        # Создаем таблицу для отображения файлов и магазинов
        self.table = QTableWidget(len(file_shop_map), 2)
        self.table.setHorizontalHeaderLabels(["Файл", "Магазин"])
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setStretchLastSection(True)
        
        # Заполняем таблицу данными
        for row, (filename, shop) in enumerate(file_shop_map.items()):
            self.table.setItem(row, 0, QTableWidgetItem(filename))
            self.table.setItem(row, 1, QTableWidgetItem(shop))
        
        layout.addWidget(self.table)
        
        # Кнопки подтверждения
        button_box = QDialogButtonBox()
        self.confirm_button = button_box.addButton("Верно", QDialogButtonBox.ButtonRole.AcceptRole)
        self.reject_button = button_box.addButton("Неверно", QDialogButtonBox.ButtonRole.RejectRole)
        
        self.confirm_button.clicked.connect(self.accept)
        self.reject_button.clicked.connect(self.reject)
            
        layout.addWidget(button_box)
        self.setLayout(layout)

class MainWindow(QMainWindow): # Класс окна для подтверждения магазинов
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Генератор заявок на закупку, лаб. РТО (alpha 0.1)")
        self.setGeometry(300, 300, 600, 600)
        self.setFixedSize(600, 600)
        

        self.label_hello = QLabel(
            "Эта утилита предназначена для генерирования файлов заявки \nна приобретение комплектующих. "
            "Скачайте файлы Excel из \nкорзины таких магазинов как 'Минимакс', 'ЭТМ', 'Чип и Дип', 'Платан', \n'Все инструменты'"
            " и загрузите их в утилиту\n \n" \
            "Автор : А.И. Сметанкин", self
        )
        self.label_hello.setAlignment(Qt.AlignmentFlag.AlignJustify)
        self.label_hello.setGeometry(40, 20, 800, 160)
        self.label_hello.setStyleSheet("font-size: 14px; font-style: italic;")

        self.button = QPushButton("Выбрать файлы Excel", self)
        self.button.setGeometry(40, 40 + 120, 150, 40)
        self.button.clicked.connect(self.process_files)

        self.label = QLabel("Выберите файлы для обработки", self)
        self.label.setGeometry(40, 100 + 120, 800, 30)
        self.label.setStyleSheet("font-size: 14px;")

        self.label_defined_shop = QLabel("", self)
        self.label_defined_shop.setGeometry(40, 140 + 120, 800, 200)
        self.label_defined_shop.setStyleSheet("font-size: 14px;")
        self.label_defined_shop.setAlignment(Qt.AlignmentFlag.AlignTop)

        # Добавляем кнопку для сохранения файла (изначально скрыта)
        self.save_button = QPushButton("Сохранить файл Excel", self)
        self.save_button.setGeometry(300, 40 + 120, 150, 40)
        self.save_button.setVisible(False)  # Сначала невидима
        self.save_button.clicked.connect(self.save_generated_file)

        # Добавляем горизонтальную линию с помощью QFrame (жирная линия)
        self.line = QFrame(self)
        self.line.setGeometry(40, 400, 520, 2)  # увеличиваем высоту до 4px
        self.line.setFrameShape(QFrame.Shape.HLine)
        self.line.setFrameShadow(QFrame.Shadow.Sunken)
        self.line.setStyleSheet("background-color: black;")  # делаем линию жирной и черной

        self.label_select_work = QLabel("Выберите наименование работы:",self)
        self.label_select_work.setGeometry(40,420,510,20)
        self.label_select_work.setStyleSheet("font-size: 14px;")

        self.work_qbox = QComboBox(self)
        self.work_qbox.addItems(work_list)
        self.work_qbox.setGeometry(40,460,510,20)
        self.work_qbox.setVisible(True)
        self.work_qbox.currentTextChanged.connect(self.text_changed)
        self.work_qbox.currentIndexChanged.connect( self.index_changed)

        self.generated_word_but = QPushButton("Создать Word", self)
        self.generated_word_but.setGeometry(40, 520, 150, 40)
        self.generated_word_but.setStyleSheet("font-size: 14px;")
        self.generated_word_but.clicked.connect(self.word_file_generated)


    def text_changed(self, s):  #выбор работы из списка
        global selected_work
        selected_work = s

    def index_changed(self, i): #присваивание статьи ДДС по индексу работы из списка
        global selected_dds
        if 0 <= i < len(dds_list):
            selected_dds = dds_list[i]


    def process_files(self):
        file_paths = ep.select_files(self)
        if not file_paths:
            return
            
        # Сначала определяем магазины для всех файлов
        file_shop_map = {}
        for file_path in file_paths:
            try:
                df = ep.read_data(file_path)
                shop_name = ep.detect_shop(file_path, df, self)
                filename = os.path.basename(file_path)
                file_shop_map[filename] = shop_name
            except Exception as e:
                print(f"Ошибка при определении магазина для {file_path}: {str(e)}")
                filename = os.path.basename(file_path)
                file_shop_map[filename] = "Ошибка определения"
        
        # Показываем пользователю таблицу соответствия
        dialog = ShopConfirmationDialog(file_shop_map, self)
        result = dialog.exec()
        
        # Если пользователь сказал "Неверно", запрашиваем магазины вручную
        if result == QDialog.DialogCode.Rejected:
            for file_path in file_paths:
                filename = os.path.basename(file_path)
                shop_name, ok = ep.get_shop_name_from_user(self, filename)
                if ok and shop_name.strip():
                    file_shop_map[filename] = shop_name
        
        # Обновляем статус в интерфейсе
        status_text = "Определённые магазины:\n"
        for filename, shop in file_shop_map.items():
            status_text += f"{filename}: {shop}\n"
        self.label_defined_shop.setText(status_text)
        QApplication.processEvents()
        
        # Теперь обрабатываем файлы с подтвержденными магазинами
        all_numbers, all_products, all_articles, all_quantities, all_costs, all_sum, all_shops = ep.process_files_with_shops(
            file_paths, self, file_shop_map
        )
        
        # Формируем результат
        df_result = pd.DataFrame({
            '№': all_numbers,
            'Наименование': all_products,
            'Магазин': all_shops,
            'Артикул': all_articles,
            'Кол-во': all_quantities,
            'Цена': all_costs,
            'Сумма': all_sum
        })
        
        # Сохраняем результат в Excel
        out_dir = "out"
        if not os.path.exists(out_dir):
            os.makedirs(out_dir)
        filename = f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(out_dir, filename)
        
        # Создаем временный файл для данных
        temp_filepath = os.path.join(out_dir, "temp_" + filename)
        df_result.to_excel(temp_filepath, index=False)
        
        try:
            # Создаем новую книгу для итогового результата
            wb_result = Workbook()
            ws_result = wb_result.active
            
            # Копируем header.xlsx если он существует
            header_path = "header.xlsx"
            header_exists = os.path.exists(header_path)
            header_row_count = 0

            
            
            if header_exists:
                wb_header = load_workbook(header_path)
                ws_header = wb_header.active
                
                # Копируем все ячейки
                for row in ws_header.iter_rows():
                        for cell in row:
                            ws_result.cell(
                                row=cell.row, 
                                column=cell.column, 
                                value=cell.value
                            )
                        
                
                # Копируем объединенные ячейки
                for merged_range in ws_header.merged_cells.ranges:
                    ws_result.merge_cells(str(merged_range))
                
                header_row_count = ws_header.max_row

                current_date = datetime.now() #получить текущую дату
                month_cell = ws_result.cell(row=6, column=6) #Записать месяц в F6
                month_cell.value = MONTHS_RU[current_date.month]

                year_cell = ws_result.cell(row=6, column=7) #Записать год в G6
                year_cell.value = f"{current_date.year} г."


            # Копируем основную таблицу с отступом
            wb_temp = load_workbook(temp_filepath)
            ws_temp = wb_temp.active
            
            start_row = header_row_count + 2 if header_exists else 1
            
            for row in ws_temp.iter_rows():
                for cell in row:
                    ws_result.cell(
                        row=cell.row + start_row - 1,
                        column=cell.column,
                        value=cell.value
                    )
            
            # Удаляем временный файл
            os.remove(temp_filepath)
            
            # Устанавливаем ширину столбцов
            ws_result.column_dimensions['A'].width = 4
            ws_result.column_dimensions['B'].width = 40
            ws_result.column_dimensions['C'].width = 18
            ws_result.column_dimensions['D'].width = 18
            ws_result.column_dimensions['E'].width = 8
            ws_result.column_dimensions['F'].width = 8
            ws_result.column_dimensions['G'].width = 8
            
            # Выравнивание данных
            center_alignment = Alignment(horizontal='center', vertical='center')
            data_start_row = start_row + 1
            data_end_row = start_row + len(df_result)
            
            # Центрирование для столбцов A, C-G
            for col_idx in [1, 3, 4, 5, 6, 7]:  # A, C, D, E, F, G
                for row_idx in range(start_row, data_end_row + 1):
                    cell = ws_result.cell(row=row_idx, column=col_idx)
                    cell.alignment = center_alignment
            
            # Перенос текста для столбца B
            for row_idx in range(data_start_row, data_end_row + 1):
                cell = ws_result.cell(row=row_idx, column=2)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Вставляем подвал из bottom.xlsx
            bottom_path = "bottom.xlsx"
            bottom_exists = os.path.exists(bottom_path)
            if bottom_exists:
                wb_bottom = load_workbook(bottom_path)
                ws_bottom = wb_bottom.active
                
                # Начальная строка для подвала: после основной таблицы + 2 пустые строки
                start_bottom_row = data_end_row + 3
                
                # Копируем ячейки
                for row in ws_bottom.iter_rows():
                        for cell in row:
                            ws_result.cell(
                                row=cell.row + start_bottom_row - 1,
                                column=cell.column,
                                value=cell.value
                            )  
                
                # Копируем объединенные ячейки
                for merged_range in ws_bottom.merged_cells.ranges:
                    # Сдвигаем диапазон
                    new_range = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row + start_bottom_row - 1}:{get_column_letter(merged_range.max_col)}{merged_range.max_row + start_bottom_row - 1}"
                    ws_result.merge_cells(new_range)
            

            # Шрифт для всех ячеек
            common_font = Font(name='Times New Roman', size=12)
            bold_font = Font(name='Times New Roman', size=12, bold=True)
    
            # Стиль границ
            thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
            )
    
            # Установка общего шрифта для всех ячеек
            for row in ws_result.iter_rows():
                for cell in row:
                    cell.font = common_font

            # Жирный шрифт для A1
            if ws_result['A1'].value:
                ws_result['A1'].font = bold_font

            # Обработка строки 9: центрирование + жирный шрифт
            if ws_result.max_row >= 9:
                for col in range(1, ws_result.max_column + 1):
                    cell = ws_result.cell(row=9, column=col)
                    cell.font = bold_font
                    # Сохраняем текущие настройки вертикального выравнивания
                    current_align = cell.alignment
                    new_align = Alignment(
                        horizontal='center',
                        vertical=current_align.vertical if current_align else 'center',
                        wrap_text=current_align.wrap_text if current_align else False
                    )
                    cell.alignment = new_align

            # Границы для основной таблицы (исключая header и bottom)
            for row_idx in range(start_row, data_end_row + 1):
                for col_idx in range(1, 8):  # Столбцы A-G
                    cell = ws_result.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    # Сохраняем итоговый файл
                    wb_result.save(filepath)
                    # Сохраняем путь к сгенерированному файлу

            self.generated_file_path = filepath
            # Показываем кнопку сохранения
            self.save_button.setVisible(True)

            
        except Exception as e:
            print(f"Ошибка при формировании файла: {str(e)}")
            # Если возникла ошибка - сохраняем без шапки и подвала
            df_result.to_excel(filepath, index=False)
            if os.path.exists(temp_filepath):
                os.remove(temp_filepath)
        
        self.label.setText(f"Результаты сохранены в файл: {filepath}")


    def save_generated_file(self):
        """Сохраняет сгенерированный файл в выбранной пользователем директории"""
        if not self.generated_file_path or not os.path.exists(self.generated_file_path):
            self.label.setText("Ошибка: файл не сгенерирован")
            return
            
        # Получаем имя файла для предложения в диалоге
        filename = os.path.basename(self.generated_file_path)
        
        # Открываем диалог выбора директории
        options = QFileDialog.Option.ShowDirsOnly
        directory = QFileDialog.getExistingDirectory(
            self,
            "Выберите папку для сохранения файла",
            options=options
        )
        
        if not directory:  # Пользователь отменил
            return
            
        # Формируем полный путь назначения
        dest_path = os.path.join(directory, filename)
        
        try:
            # Копируем файл
            import shutil
            shutil.copy(self.generated_file_path, dest_path)
            self.label.setText(f"Файл успешно сохранён в:\n{dest_path}")
        except Exception as e:
            self.label.setText(f"Ошибка при сохранении: {str(e)}")
    

    def word_file_generated(self):
        from datetime import datetime
        try:
            # Получаем текущую дату для подстановки
            current_date = datetime.now().strftime("%d.%m.%Y")
            
            # Генерируем Word-файл
            generated_file_path = wp.generated_wfile(
                doc=document,
                substring_map=sub_map, 
                work_name=selected_work, 
                dds=selected_dds,
                date_str=current_date
            )
            
            # Показываем сообщение о успешной генерации
            self.label.setText(f"Word файл сгенерирован:\n{generated_file_path}")
        except Exception as e:
            self.label.setText(f"Ошибка при генерации Word: {str(e)}")
            print(f"Ошибка: {str(e)}")

if __name__ == "__main__":  
    work_list_filepath = "work_list.xlsx"
    data_from_excel = ep.read_data(work_list_filepath)
    work_list = data_from_excel.iloc[:, 0].tolist()
    selected_work = work_list[0] if work_list else ""

    dds_list = data_from_excel.iloc[:, 1].tolist()
    selected_dds = dds_list[0] if dds_list else ""

    # Загружаем документ только если файл существует
    if os.path.exists("sample.docx"):
        document = Document("sample.docx")
        sub_map = wp.searc_text_position(document)
    else:
        document = None
        sub_map = {}
        print("Предупреждение: файл sample.docx не найден")

    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())